import openpyxl
import json
import hashlib

participant_path = r'c:\Users\Sergio\Documents\cosas_antigravity\porra_mundial\Porra_Mundial2026_Automatizacion_Total.xlsx'
db_path = r'c:\Users\Sergio\Documents\cosas_antigravity\porra_mundial\data.json'

print("Loading Excel workbooks for seeding...")
wb_formulas = openpyxl.load_workbook(participant_path, data_only=False)
wb_vals = openpyxl.load_workbook(participant_path, data_only=True)
sheet_calc_formulas = wb_formulas['Calculos_Internos']
sheet_calc_vals = wb_vals['Calculos_Internos']

def resolve_ref(ref_str, wb_vals):
    if not ref_str or not str(ref_str).startswith('='):
        return ref_str
    ref = str(ref_str).replace('=', '').replace('$', '')
    if '!' in ref:
        s_name, cell_c = ref.split('!')
        s_name = s_name.strip("'")
        return wb_vals[s_name][cell_c].value
    return ref_str

matches = []

# 1. Read group stage matches (Rows 2 to 73)
print("Reading group stage matches...")
for r in range(2, 74):
    match_id = sheet_calc_vals.cell(row=r, column=1).value
    fase = sheet_calc_vals.cell(row=r, column=2).value
    local_formula = sheet_calc_formulas.cell(row=r, column=3).value
    visitor_formula = sheet_calc_formulas.cell(row=r, column=6).value
    
    local = resolve_ref(local_formula, wb_vals)
    visitor = resolve_ref(visitor_formula, wb_vals)
    
    matches.append({
        "id": match_id,
        "group": fase.replace("Grupo ", ""),
        "phase": "Group Stage",
        "local": local,
        "visitor": visitor,
        "gl": None,
        "gv": None,
        "pkl": None,
        "pkv": None
    })

# 2. Read knockout stage matches (Rows 200 to 231)
print("Reading knockout stage matches...")
for r in range(200, 232):
    match_id = sheet_calc_vals.cell(row=r, column=1).value
    fase = sheet_calc_vals.cell(row=r, column=2).value
    
    matches.append({
        "id": match_id,
        "group": None,
        "phase": fase,
        "local": f"Local {match_id}",
        "visitor": f"Visitante {match_id}",
        "gl": None,
        "gv": None,
        "pkl": None,
        "pkv": None
    })

# Define default admin credentials
# salt = "porrasecret"
# password = "admin123"
# hash = sha256(password + salt)
salt = "porrasecret"
admin_pass = "admin123"
admin_hash = hashlib.sha256((admin_pass + salt).encode()).hexdigest()

db_initial = {
    "config": {
        "points": {
            "outcome": 1,
            "exact": 3,
            "balon_oro": 10,
            "balon_plata": 5,
            "balon_bronce": 3,
            "bota_oro": 10,
            "bota_plata": 5,
            "bota_bronce": 3
        },
        "winners": {
            "balon_oro": "",
            "balon_plata": "",
            "balon_bronce": "",
            "bota_oro": "",
            "bota_plata": "",
            "bota_bronce": ""
        }
    },
    "users": [
        {
            "id": "admin_user",
            "username": "admin",
            "passwordHash": admin_hash,
            "salt": salt,
            "isAdmin": True
        }
    ],
    "predictions": {},
    "matches": matches,
    "rankingHistory": []
}

print(f"Saving database JSON to {db_path}...")
with open(db_path, 'w', encoding='utf-8') as f:
    json.dump(db_initial, f, indent=2, ensure_ascii=False)

print("Database seeding completed successfully!")
